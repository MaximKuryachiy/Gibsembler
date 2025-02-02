import re
import sys
import subprocess
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from Bio import SeqIO
from Bio.SeqUtils import MeltingTemp as mt
import importlib.util

# Ensure missing libraries are installed
def install_missing_libraries():
    required_libraries = ["pandas", "openpyxl", "biopython"]
    missing_libraries = [lib for lib in required_libraries if not importlib.util.find_spec(lib)]
    
    if missing_libraries:
        for lib in missing_libraries:
            subprocess.check_call([sys.executable, "-m", "pip", "install", lib])
        print("Libraries installed. Restarting the program.")
        
    if missing_libraries:
        print("Libraries installed. Please restart the program manually.")
        return
python_exe = sys.executable

install_missing_libraries()

# DNA analysis functions
def calculate_gc_content(sequence):
    return (sequence.count("G") + sequence.count("C")) / len(sequence) * 100 if sequence else 0

def calculate_tm(sequence):
    return mt.Tm_NN(sequence) if sequence else 0

def read_sequence(file_path):
    print(f"Reading sequence from: {file_path}")
    file_extension = file_path.split(".")[-1].lower()
    try:
        record = SeqIO.read(file_path, "genbank" if file_extension in ["gb", "gbk"] else "fasta")
        sequence = str(record.seq)
        print(f"Sequence length: {len(sequence)}")
        return sequence
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read file: {e}")
        return ""

def optimize_segment(dna_seq, start, min_length=20, max_length=50, target_gc=(40, 60), target_tm=(57, 60)):
    best_segment, best_gc, best_tm, best_length = None, None, None, None
    for length in range(min_length, max_length + 1):
        end = (start + length) % len(dna_seq)
        segment = dna_seq[start:end] if end > start else dna_seq[start:] + dna_seq[:end]
        gc, tm = calculate_gc_content(segment), calculate_tm(segment)
        if target_gc[0] <= gc <= target_gc[1] and target_tm[0] <= tm <= target_tm[1]:
            return segment, length, gc, tm
        if best_tm is None or abs(tm - sum(target_tm) / 2) < abs(best_tm - sum(target_tm) / 2):
            best_segment, best_gc, best_tm, best_length = segment, gc, tm, length
    return best_segment, best_length, best_gc, best_tm

def segment_dna_for_gibson(dna_seq, min_length=20, max_length=50, overlap_min=10, overlap_max=15, max_iterations=100000):
    print("Starting segmentation...")
    segments, i, seq_length = [], 0, len(dna_seq)
    iteration_count = 0
    visited_positions = set()
    while i < seq_length and iteration_count < max_iterations and i not in visited_positions:
        segment, segment_length, gc, tm = optimize_segment(dna_seq, i, min_length, max_length)
        end_index = (i + segment_length) % seq_length
        overlap_length = min(overlap_max, max(overlap_min, segment_length // 2))
        overlap = dna_seq[end_index - overlap_length:end_index] if end_index - overlap_length >= 0 else dna_seq[end_index - overlap_length:] + dna_seq[:end_index]
        overlap = overlap[:overlap_max]  # Ensure strict max overlap length
        segments.append((segment, segment_length, gc, tm, overlap, overlap_length))
        visited_positions.add(i)
        i = (i + segment_length - overlap_length) % seq_length
        iteration_count += 1
        if iteration_count % 1000 == 0:
            print(f"Processing... Iteration {iteration_count}, Current Index: {i}")
            print(f"Processed {len(segments)} segments.")
    if iteration_count >= max_iterations:
        print("Warning: Maximum iterations reached. Possible infinite loop detected.")
    return segments

def save_segments(output_file, segments, file_format):
    print(f"Saving {len(segments)} segments to {output_file} in {file_format} format...")
    if file_format == "Excel":
        wb = Workbook()
        ws = wb.active
        ws.append(["Segment Sequence", "Length", "GC Content (%)", "Tm (°C)", "Overlap Sequence", "Overlap Length"])
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        for row_idx, (segment, length, gc, tm, overlap, overlap_length) in enumerate(segments, start=2):
            ws.append([segment, length, gc, tm, overlap, overlap_length])
            for col_idx, value in enumerate([segment, overlap], start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if overlap in str(value):  
                    cell.fill = red_fill  
        wb.save(output_file)
    else:
        df = pd.DataFrame(segments, columns=["Segment Sequence", "Length", "GC Content (%)", "Tm (°C)", "Overlap Sequence", "Overlap Length"])
        df.to_csv(output_file, index=False)
    print(f"Segments saved to {output_file}")

def get_parameters():
    param_window = tk.Toplevel()
    param_window.title("Segmentation Parameters")
    tk.Label(param_window, text="Target GC Content (Min-Max):").grid(row=0, column=0)
    gc_entry = tk.Entry(param_window)
    gc_entry.insert(0, "40-60")
    gc_entry.grid(row=0, column=1)
    tk.Label(param_window, text="Target Tm (Min-Max):").grid(row=1, column=0)
    tm_entry = tk.Entry(param_window)
    tm_entry.insert(0, "57-60")
    tm_entry.grid(row=1, column=1)
    tk.Label(param_window, text="Overlap Length (Min-Max):").grid(row=2, column=0)
    overlap_entry = tk.Entry(param_window)
    overlap_entry.insert(0, "10-15")
    overlap_entry.grid(row=2, column=1)
    def save_parameters():
        param_window.destroy()
        return [gc_entry.get(), tm_entry.get(), overlap_entry.get()]
    tk.Button(param_window, text="OK", command=save_parameters).grid(row=3, columnspan=2)
    param_window.mainloop()

def main():
    root = tk.Tk()
    root.withdraw()
    input_file = filedialog.askopenfilename(title="Select DNA Sequence File", filetypes=[("FASTA or GenBank Files", "*.fasta;*.gb;*.gbk")])
    if not input_file:
        return
    save_path = filedialog.asksaveasfilename(title="Save File As", defaultextension=".csv", filetypes=[("CSV File", "*.csv"), ("Excel File", "*.xlsx")])
    if not save_path:
        return
    file_format = "Excel" if save_path.endswith(".xlsx") else "CSV"
    dna_sequence = read_sequence(input_file)
    gibson_segments = segment_dna_for_gibson(dna_sequence)
    save_segments(save_path, gibson_segments, file_format)
    messagebox.showinfo("Success", f"Segments saved successfully as {save_path}")

if __name__ == "__main__":
    main()
